import openpyxl
import datetime

from openpyxl import load_workbook

caminho_arquivo = "C:/Users/Quixabeira/Documents/Projetos/excel-python/dados.xlsx"
workbook = load_workbook(caminho_arquivo)

# Selecionar uma planilha (por nome ou ativa)
worksheet =   workbook['VENDAS'] # ou workbook.active

line_count = 0

data_vendas = []

# Ler valores das linhas
for line in worksheet:

    # Obter os valores das células, ignorando None
    line = [cell.value for cell in line if cell.value is not None]

    line_info = []

    line_count = line_count + 1 #conta as linhas
    cell_count = 0
    
    if line != [] and line_count > 1: #ignora lists vazias e a primeira linha

        for cell in line: #ler os valores das celulas e agrupalas em uma lista
            cell_count = cell_count + 1

            if cell_count >= 1:
                line_info.append(cell) 

        if len(line_info) > 1:
            data_vendas.append(line_info)

workbook.close()

#A variável //data// contem uma Lista com listas dos dados de cada linha, o primeiro valor é a linha dessa dado no excel, ex:
#[2, 'Samsung Galaxy S23 Ultra', 10, 9, 12, 5] -> Nesse exemplo, essa dado está na linha 2 no excel

#filtrando os dados -------------------------------------

for index,i in enumerate(data_vendas):
    id = i[0]
    name = i[1]  # Nome do produto
    sum_all = sum(i[2:])  # Soma dos valores a partir do índice 2

    data_vendas[index] = {'id':id,
         'product':name,
         'sells':sum_all}
    
#Tudo foi separado em dicionários por pura organização

#Printando os dados -------------------------------------

#for i in data_vendas:
 #  print(i)

#Mesclando dados com a tabela "valor"--------------------
workbook = load_workbook(caminho_arquivo)

worksheet =   workbook['VALOR']

line_count = 0
data_valor = []

# Ler valores das linhas
for line in worksheet:

    # Obter os valores das células, ignorando None
    line = [cell.value for cell in line if cell.value is not None]

    line_count = line_count + 1 #conta as linhas
    cell_count = 0
    line_info = []
    
    if line != [] and line_count > 1: #ignora lists vazias e a primeira linha

        for cell in line: #ler os valores das celulas e agrupalas em uma lista
            cell_count = cell_count + 1

            if cell_count >= 1:
                line_info.append(cell) 

        if len(line_info) > 1:
            data_valor.append(line_info)

for index,i in enumerate(data_valor):
    id = i[0]
    name = i[1]  # Nome do produto
    price = i[2]  # preço de venda
    acsition = float(price) * 0.65

    data_valor[index] = {'id':id,
         'product':name,
         'price':price,
         'acsition':acsition}
    
#Printando os dados -------------------------------------

#print("")

#for i in data_valor:
 #   print(i)

#Analisando ambos os dados -------------------------------------

for vendas, valor in zip(data_vendas, data_valor):
    id = vendas.get('id')
    name = vendas.get('product')
    sells = vendas.get('sells')
    price = "R${:,.2f}".format(valor.get('price') * sells)
    acsition = "R${:,.2f}".format(valor.get('acsition') * sells)
    profit = "R${:,.2f}".format(valor.get('price') * sells - valor.get('acsition') * sells)

    print(f"ID: {id} \nName: {name} \nSells quantity: {sells}\nTotal sold: {price} \nAcsition price: {acsition}\nProfit: {profit}\n")